"""Engine orchestrator — the single entry point the trigger service calls.

    build_bom(planset_pdf_path, coperniq_project_dict) -> (xlsx_bytes, confidence_dict)

This wraps the validated computed engine (racking_engine, filter_blank_rows) and the template fill.
The EXTRACTION half (reading orientation/dims/breakers/disconnects off the planset) is the part still
being validated for autonomy — it lives behind `extract_planset()` and currently raises
NeedsHumanExtraction unless an extractor is wired. That deliberate failure routes through the
pipeline's "generation failed, needs human" path rather than shipping a guessed BOM.

To go live in shadow mode you wire extract_planset() to the planset-extractor (Claude Vision) and the
confidence gates from AUTONOMY_READINESS_SPEC. Until then, the service is fully functional end-to-end
EXCEPT this boundary, which is exactly the validated/unvalidated line we documented.
"""
from __future__ import annotations
import io
import math
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from . import racking_engine as re_eng
from .filter_blank_rows import apply_qty_filter

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(HERE, "BOM_TEMPLATE.xlsx")
LIBREOFFICE_BIN = os.environ.get("LIBREOFFICE_BIN", "libreoffice")


class NeedsHumanExtraction(RuntimeError):
    """Raised when automated extraction is not wired/validated for this planset. Routes to the
    pipeline's failure-notify path so a human builds the BOM instead of shipping a guess."""


# ---------- extracted-planset data model ----------
@dataclass
class ExtractedArray:
    label: str
    azimuth: float
    rail_runs: list[int]                 # module count per rail run (rows), left->right segments
    orientation: str = "landscape"       # uniform per run for now; mixed handled via per-run lists

@dataclass
class Extracted:
    module_long_in: float
    module_short_in: float
    module_sku_row: int                  # Solar row 5-9 by module model
    module_count: int
    mci_count: int
    arrays: list[ExtractedArray]
    planset_racking: dict                 # attachments, rails, splice, mid_clamps, end_clamps
    attachment_type: str                  # K2_SHINGLE | S5_PROTEABRACKET | S5_SOLARFOOT
    roof_types: list[str]                 # per array: shingle|metal|ground
    # electrical
    ac_disco: dict                        # {fused: bool, rating: int}
    dc_strings: int
    battery_pw3: int
    expansion: int
    gateway: bool
    gateway_buskit_60a: int
    gateway_breakers_outside_buskit: list[int]
    meter_sku_row: Optional[int]
    new_meter: bool
    flags: list[dict] = field(default_factory=list)


def extract_planset(planset_pdf_path: str, project: dict) -> Extracted:
    """EXTRACTION BOUNDARY (still being validated — see AUTONOMY_READINESS_SPEC §A).

    Wire this to the planset-extractor (Claude Vision over PV-1/PV-3/PV-3.1/PV-5) + the Coperniq
    cross-check. It MUST populate module dims from PV-3, per-array orientation from the rotated
    raster, the planset racking table, attachment type, and the electrical reads — appending a flag
    (level HARD/SOFT/INFO) for every gate in the spec that fires.

    Until wired, raise NeedsHumanExtraction so the pipeline notifies a human rather than guessing.
    """
    raise NeedsHumanExtraction(
        "Automated planset extraction is not yet validated for autonomous use. Wire "
        "engine/orchestrator.extract_planset() to the planset-extractor + confidence gates "
        "(AUTONOMY_READINESS_SPEC §A) before enabling unattended runs."
    )


# ---------- module-row map ----------
def _module_row(sku_hint: str) -> int:
    s = (sku_hint or "").upper()
    if "ELNSM54M" in s: return 5
    if "Q.PEAK" in s or "QCELL" in s: return 6
    if "HIS-T440" in s or "HYUNDAI" in s: return 7
    if "DNA-120" in s or "APTOS" in s: return 8
    if "LR5-54HPB" in s or "LONGI" in s: return 9
    return 5


# ---------- build the workbook from extracted data + engine ----------
def _fill_and_recalc(ext: Extracted, project: dict, confidence: dict) -> bytes:
    md = re_eng.ModuleDims.from_pv3(ext.module_long_in, ext.module_short_in)

    # build engine rows (orientation already from rotated raster in extraction)
    arrays = []
    for a in ext.arrays:
        rows = [re_eng.make_row(n, a.orientation, round(n * md.edge(a.orientation), 1),
                                from_rotated_raster=True, long_in=md.long_in, short_in=md.short_in)
                for n in a.rail_runs]
        arrays.append({"label": a.label, "azimuth": a.azimuth, "rows": rows})

    delivered, xc = re_eng.resolve_racking(arrays, ext.planset_racking,
                                           attachment_type=ext.attachment_type, module_dims=md)
    confidence["racking_crosscheck"] = xc

    # gateway breakers (CSR rule)
    gw = re_eng.tesla_gateway_breakers(ext.gateway_buskit_60a, ext.battery_pw3,
                                       ext.gateway_breakers_outside_buskit) if ext.gateway else {}

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    S, E = wb["Solar BOM"], wb["Electrical BOM"]

    # clear numeric col-A (skip A30 S-clip)
    for ws in (S, E):
        for r in range(5, ws.max_row + 1):
            if ws is S and r == 30:
                continue
            if isinstance(ws.cell(r, 1).value, (int, float)):
                ws.cell(r, 1).value = None

    # headers
    addr = project.get("address")
    S["B1"] = project.get("title", "")
    z = (project.get("custom", {}) or {}).get("zone", [""])
    S["B2"] = z[0] if isinstance(z, list) and z else (z or "")
    S["B3"] = addr[0] if isinstance(addr, list) and addr else (addr or "")

    def setq(ws, r, q):
        if q:
            ws.cell(r, 1).value = q

    # ---- Solar ----
    setq(S, ext.module_sku_row, ext.module_count)
    setq(S, 20, ext.mci_count)
    # J-box by roof type (metal->26, shingle->25); ground always 26
    # (extraction supplies roof_types + strings; here simplified to delivered split)
    # racking
    if ext.attachment_type == "S5_PROTEABRACKET":
        setq(S, 37, ext.planset_racking["attachments"])
    elif ext.attachment_type == "S5_SOLARFOOT":
        setq(S, 38, ext.planset_racking["attachments"])
        setq(S, 39, ext.planset_racking["attachments"])  # separate L-foot
    else:  # K2 shingle
        setq(S, 33, ext.planset_racking["attachments"])
        setq(S, 35, delivered["deck_screw"])             # 4x attach
    setq(S, 45, delivered["rails"])
    setq(S, 46, delivered["rail_clamp"])
    setq(S, 47, delivered["splice"])
    setq(S, 48, delivered["combo_clamp"])
    setq(S, 50, delivered["ground_lug"])
    setq(S, 51, delivered["end_cap"])
    setq(S, 52, delivered["wire_clip"])

    # ---- Electrical ----
    if ext.ac_disco:
        disco_rows_nf = {30: 5, 60: 6, 100: 7, 200: 8}
        disco_rows_f = {60: 10, 100: 11, 200: 12}
        rating = ext.ac_disco.get("rating", 60)
        row = (disco_rows_f if ext.ac_disco.get("fused") else disco_rows_nf).get(rating)
        if row:
            setq(E, row, 1)
            # hub: 2x per disco by size bucket
            if rating <= 60:
                setq(E, 13, 2)
            elif rating == 100:
                setq(E, 14, 2)
            else:
                setq(E, 15, 2)
    if ext.battery_pw3:
        setq(E, 19, math.ceil(ext.battery_pw3 / 3))      # RSD PE69-3020
        setq(E, 52, ext.battery_pw3)                      # PW3 domestic default
    if ext.gateway:
        setq(E, 22, 1)                                    # ground bar per gateway
        setq(E, 54, 1)
    if ext.expansion:
        setq(E, 59, ext.expansion)
        setq(E, 62, ext.expansion)                        # wall-mount default
        setq(E, 64, ext.expansion)                        # -20 harness
    if ext.new_meter and ext.meter_sku_row:
        setq(E, ext.meter_sku_row, 1)
    for row, q in gw.items():
        if isinstance(row, int):
            setq(E, row, q)

    # save -> recalc -> filter
    tmpdir = tempfile.mkdtemp()
    raw_path = os.path.join(tmpdir, "bom_raw.xlsx")
    wb.save(raw_path)
    try:
        subprocess.run([LIBREOFFICE_BIN, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", tmpdir, raw_path], check=True, timeout=180,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        # recalc failed; flag it HARD and fall back to the unrecalculated file
        confidence.setdefault("FLAGS_FOR_HUMAN_REVIEW", []).append(
            {"level": "HARD", "item": "recalc_failed",
             "msg": "LibreOffice recalc failed; A30 S-clip may be stale. Open + recalc manually."})
    recalc_path = raw_path  # convert-to writes same name in outdir
    apply_qty_filter(recalc_path)
    with open(recalc_path, "rb") as f:
        data = f.read()
    shutil.rmtree(tmpdir, ignore_errors=True)
    return data


def build_bom(planset_pdf_path: str, coperniq_project_dict: dict) -> tuple[bytes, dict]:
    """Public entry point for the trigger service."""
    confidence: dict = {
        "project": {
            "id": coperniq_project_dict.get("id"),
            "name": coperniq_project_dict.get("title"),
            "number": coperniq_project_dict.get("number"),
        },
        "FLAGS_FOR_HUMAN_REVIEW": [],
        "mode": "shadow",
    }
    ext = extract_planset(planset_pdf_path, coperniq_project_dict)  # raises until wired
    # SECOND-STAGE PLANSET CONFIRMATION: once extraction reads PV-1 text, verify it matches this
    # project (catches a correctly-NAMED file containing the wrong project's plans). The extractor
    # should attach these via confirm_planset_content(pv1_text, name, address); they land as flags.
    confidence["FLAGS_FOR_HUMAN_REVIEW"].extend(ext.flags)
    xlsx_bytes = _fill_and_recalc(ext, coperniq_project_dict, confidence)
    return xlsx_bytes, confidence
