"""Canonical blank-row filter for Tron BOM output.

RULE: in the data region (row 5 down), hide EVERY row whose QUANTITY (col A) is empty/blank/zero
-- menu SKU rows OR fully-empty spacer rows. Exempt: a formula row (e.g. Solar A30 S-clip
'=IF(...)') is kept if it holds a formula.

MECHANISM: a real Excel AUTOFILTER anchored on the FULL table range (A:D), plus the matching row
.hidden flags. Excel persists both -- the filter definition and the rows it hides. Because it's a
real AutoFilter, the user can clear/toggle it in Excel and ALL hidden rows reappear; the hidden
rows stay in the sheet, just filtered out. Applies to 'Solar BOM' and 'Electrical BOM'.
"""
import openpyxl

DATA_START = 5  # data begins at row 5 (row 4 = header)


def _has_qty(ws, r):
    """True if col A holds a real quantity (numeric != 0, or a formula string)."""
    v = ws.cell(r, 1).value
    if v is None:
        return False
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return False
        if s.startswith("="):   # formula (e.g. S-clip) counts as present
            return True
        try:
            return float(s) != 0
        except ValueError:
            return False
    try:
        return float(v) != 0
    except (TypeError, ValueError):
        return False


def apply_qty_filter(path, table_last_col="D"):
    """Hide every data row with no/zero qty via a real AutoFilter over the full A:<last> table."""
    wb = openpyxl.load_workbook(path)
    for sh in ("Solar BOM", "Electrical BOM"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        last_row = ws.max_row
        # 1) AutoFilter anchored on the FULL table range so clearing it reveals whole rows.
        ws.auto_filter.ref = f"A4:{table_last_col}{last_row}"
        ws.auto_filter.filterColumn = []  # reset any stale value-list criteria
        # 2) Per-row hidden flags: hide data rows without qty; keep header + rows with qty.
        for r in range(DATA_START, last_row + 1):
            ws.row_dimensions[r].hidden = not _has_qty(ws, r)
    wb.save(path)


if __name__ == "__main__":
    import sys
    apply_qty_filter(sys.argv[1])
    print("filtered", sys.argv[1])
