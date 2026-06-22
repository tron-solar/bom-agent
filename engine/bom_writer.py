"""
bom_writer.py — THE canonical, hardcoded BOM sheet writer for the Tron Solar engine.

WHY THIS EXISTS (user, Carmen Meyer #877571): sheet-writing was happening in per-project ad-hoc
scripts. That is the exact failure mode the standing discipline forbids — a change agreed in one
project (e.g. stamping the special-order P/N into the line) would silently NOT happen on the next
headless run. This module makes every write step a hardcoded engine function so EVERY future project
fills BOM_TEMPLATE.xlsx identically, with NO runtime inference and NO hand-built script.

What it owns (all the mechanics that must be identical every run):
  1. write_bom()    — fill header + static quantities into Solar BOM / Electrical BOM
  2. special-order  — stamp the ACTUAL P/N into special-order lines (meter row 96, MSP row 79, …)
  3. STATIC numbers only — never formula cells (Excel Protected View blanks uncached formulas)
  4. blank-row filter via filter_blank_rows.apply_qty_filter (real AutoFilter over A:D)
  5. output naming  — BOM_First_Last.xlsx in the outputs dir

Column layout (both sheets): A=QTY, B=SKU, C=ITEM DESCRIPTION, D=BOM tag. Header: B1 customer,
B2 warehouse zone, B3 address (Electrical mirrors via ='Solar BOM'!Bn). Data starts row 5.

Inputs are the {row: qty} dicts the engine blocks already return, plus a {row: text} special-order
override map (the third element of meter_socket()/main_service_panel()). The writer never decides
quantities — it only places what the engine computed.
"""
from __future__ import annotations
import os
import shutil
from openpyxl import load_workbook

from . import filter_blank_rows  # package-relative (was bare `import filter_blank_rows` in the flat handoff)

QTY_COL = 1   # A
SKU_COL = 2   # B
DESC_COL = 3  # C
DATA_START = 5


def _stamp_special_orders(ws, special_order):
    """special_order: {row: "text"} — write the verbatim P/N text into the special-order line so it
    names the exact part to procure. The template ships these rows with a placeholder in the SKU cell
    (col B: "Special Order Meter: (Part Number)" / "Special Order MSP: (enter sku)"). We replace BOTH
    the SKU cell (col B) and the DESCRIPTION cell (col C) with the stamped text (user, Meyer #877571),
    so the placeholder never survives into the delivered BOM and the line shows the SKU in column B."""
    for row, text in (special_order or {}).items():
        ws.cell(int(row), SKU_COL).value = str(text)
        ws.cell(int(row), DESC_COL).value = str(text)


def write_bom(template_path, out_path, *, customer_name, warehouse_zone, customer_address,
              solar_rows, electrical_rows,
              solar_special_order=None, electrical_special_order=None,
              apply_filter=True):
    """Fill the template and write the final BOM. Returns out_path.

    customer_name / warehouse_zone / customer_address : header values (planset name of record for
        customer; warehouse_zone from electrical_engine.warehouse_zone(); address from planset).
    solar_rows / electrical_rows       : {row_int: qty_int} from the engine blocks (STATIC numbers).
    solar/electrical_special_order     : {row_int: "verbatim text"} to stamp into special-order lines
        (e.g. electrical_special_order = {96: "Special Order Meter: U3358-O-KK"}).
    apply_filter                       : run the canonical blank-row AutoFilter (default True).

    STATIC quantities only — never a formula string. Header B1/B2/B3 on Solar BOM; Electrical
    mirrors via the template's existing cross-sheet refs (do not overwrite Electrical B1-B3).
    """
    shutil.copyfile(template_path, out_path)
    wb = load_workbook(out_path)
    solar = wb["Solar BOM"]
    elec = wb["Electrical BOM"]

    # header — Solar BOM only; Electrical B1-B3 already ='Solar BOM'!Bn in the template
    solar["B1"] = customer_name or ""
    solar["B2"] = warehouse_zone or ""
    solar["B3"] = customer_address or ""

    # static quantities
    for r, q in (solar_rows or {}).items():
        solar.cell(int(r), QTY_COL).value = int(q)
    for r, q in (electrical_rows or {}).items():
        elec.cell(int(r), QTY_COL).value = int(q)

    # special-order P/N stamping (meter row 96, MSP row 79, any future special-order line)
    _stamp_special_orders(solar, solar_special_order)
    _stamp_special_orders(elec, electrical_special_order)

    wb.save(out_path)

    if apply_filter:
        # canonical blank-row filter: hide every data row with empty/zero qty via a real AutoFilter
        filter_blank_rows.apply_qty_filter(out_path, table_last_col="D")

    return out_path


def output_filename(first, last, outputs_dir="/mnt/user-data/outputs"):
    """Canonical output naming: BOM_First_Last.xlsx in the outputs dir."""
    safe = lambda s: "".join(c for c in str(s).strip() if c.isalnum() or c in ("-", "_")) or "Unknown"
    return os.path.join(outputs_dir, f"BOM_{safe(first)}_{safe(last)}.xlsx")


def merge_block(target, block_rows):
    """Merge a {row: qty} block into the target accumulator, summing collisions. Engine blocks each
    return their own {row:qty}; the consolidated builder merges them all before write_bom()."""
    for r, q in (block_rows or {}).items():
        target[r] = target.get(r, 0) + int(q)
    return target


def merge_special(target, block_special):
    """Merge a {row: text} special-order override map into the accumulator (last-writer-wins)."""
    for r, t in (block_special or {}).items():
        target[r] = t
    return target
